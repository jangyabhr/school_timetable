"""
generate_from_excel.py
Reads timetable_with_4_periods_v2.xlsx (All Sections + All Teachers sheets)
and writes Timetable_Tools.html using the existing html_exporter._build_html.
"""

import sys
import json
from pathlib import Path
import openpyxl
import html_exporter as hx

# ── Override module-level constants for 4-period schedule ─────────────────────
DAYS_LONG = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
DAYS      = ['Mon','Tue','Wed','Thu','Fri','Sat']
DAY_MAP   = dict(zip(DAYS_LONG, DAYS))

PERIODS   = ['P1','P2','P3','P4']
TIMES     = ['7:10–7:50','7:50–8:30','8:40–9:20','9:20–10:00']

CLASS_ORDER = ["12","10B","10A","9B","9A","8B","8A","7B","7A"]

SPECIALS = {'Free','Duty','Principal Sir'}

# Patch module constants so _build_html uses 4 periods
hx.DAYS    = DAYS
hx.PERIODS = PERIODS
hx.TIMES   = TIMES

# Patch CLASS_ORDER in both modules
import event_generator as eg
eg.CLASS_ORDER = CLASS_ORDER
hx_class_order_ref = CLASS_ORDER  # used below in json dump

SUBJ_COLORS = hx.SUBJ_COLORS
SUBJ_COLORS.update({
    'Free':         '#ECECEC',
    'Duty':         '#D5D8DC',
    'Principal Sir':'#D5D8DC',
})

# ── Parse Excel ───────────────────────────────────────────────────────────────

def clean(val):
    if val is None:
        return None
    s = str(val).split('\n')[0].strip()
    return s if s else None

def parse_sections(wb):
    """Returns {cls: {day: [slot0, slot1, slot2, slot3]}} where slot = (subject, teacher, is_lab)"""
    ws = wb['All Sections']
    sections = {}
    current_cls = None
    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        first = str(vals[0])
        if first.startswith('Timetable — Class'):
            current_cls = first.replace('Timetable — Class ','').strip()
            sections[current_cls] = {}
        elif first.startswith('ACADEMIC') or first == 'Day':
            continue
        elif current_cls and first in DAYS_LONG:
            day = DAY_MAP[first]
            # cols after filtering None: [Day, P1, P2, P3, P4]
            cells = vals[1:5]
            slots = []
            for i in range(4):
                raw = cells[i] if i < len(cells) else None
                subj_raw = clean(raw)
                if subj_raw is None:
                    slots.append(('', '', False))
                    continue
                # teacher is on second line of cell
                teacher_raw = ''
                if raw and '\n' in str(raw):
                    lines = str(raw).split('\n')
                    teacher_raw = lines[1].strip() if len(lines) > 1 else ''
                is_lab = '(Lab)' in str(raw)
                subj = subj_raw.replace(' (Lab)','').strip()
                slots.append((subj, teacher_raw, is_lab))
            sections[current_cls][day] = slots
    return sections

def parse_teachers(wb):
    """Returns {teacher: {day: [slot0..3]}} where slot = (subject, cls, is_lab) or None"""
    ws = wb['All Teachers']
    teachers = {}
    current_t = None
    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        first = str(vals[0])
        if first.startswith('Timetable —'):
            current_t = first.replace('Timetable — ','').strip()
            teachers[current_t] = {}
        elif first.startswith('ACADEMIC') or first == 'Day':
            continue
        elif current_t and first in DAYS_LONG:
            day = DAY_MAP[first]
            cells = vals[1:5]
            slots = []
            for i in range(4):
                raw = cells[i] if i < len(cells) else None
                subj_raw = clean(raw)
                if subj_raw is None:
                    slots.append(None)
                    continue
                cls_raw = ''
                if raw and '\n' in str(raw):
                    lines = str(raw).split('\n')
                    cls_raw = lines[1].strip().strip('()') if len(lines) > 1 else ''
                is_lab = '(Lab)' in str(raw)
                subj = subj_raw.replace(' (Lab)','').strip()
                slots.append((subj, cls_raw, is_lab))
            teachers[current_t][day] = slots
    return teachers

# ── Build data structures for _build_html ────────────────────────────────────

def build_structures(sections, teacher_data):
    n = len(PERIODS)

    # class_timetable[cls][day] = list of {s, t, l}
    class_timetable = {
        cls: {day: [{'s':'','t':'','l':False} for _ in range(n)] for day in DAYS}
        for cls in CLASS_ORDER
    }
    for cls in CLASS_ORDER:
        if cls not in sections:
            continue
        for day in DAYS:
            if day not in sections[cls]:
                continue
            for pi, (subj, teacher, is_lab) in enumerate(sections[cls][day]):
                class_timetable[cls][day][pi] = {'s': subj, 't': teacher, 'l': is_lab}

    all_teachers = sorted(teacher_data.keys())

    # teacher_sched[t][day][pi] = {'class': cls, 'subject': subj} or None
    # teacher_view[t][day][pi]  = {'cls': cls, 's': subj, 'l': bool} or None
    teacher_sched = {t: {day: [None]*n for day in DAYS} for t in all_teachers}
    teacher_view  = {t: {day: [None]*n for day in DAYS} for t in all_teachers}

    for t, t_days in teacher_data.items():
        for day, slots in t_days.items():
            for pi, slot in enumerate(slots):
                if slot is None:
                    continue
                subj, cls, is_lab = slot
                if subj in ('', None):
                    continue
                teacher_sched[t][day][pi] = {'class': cls, 'subject': subj}
                teacher_view[t][day][pi]  = {'cls': cls, 's': subj, 'l': is_lab}

    # wk_load
    wk_load = {t: 0 for t in all_teachers}
    for t, t_days in teacher_data.items():
        for day, slots in t_days.items():
            for slot in slots:
                if slot and slot[0] not in ('', None) and slot[0] not in SPECIALS:
                    wk_load[t] += 1

    # special_counts
    special_counts = {cls: {s: 0 for s in ('Free','Library')} for cls in CLASS_ORDER}
    for cls in CLASS_ORDER:
        if cls not in sections:
            continue
        for day in DAYS:
            if day not in sections[cls]:
                continue
            for subj, _, _ in sections[cls][day]:
                if subj in special_counts[cls]:
                    special_counts[cls][subj] += 1

    # coverage (substitute finder) — slots where class has Free/Duty
    coverage = {}
    for cls in CLASS_ORDER:
        coverage[cls] = {}
        for day in DAYS:
            cov_list = []
            for pi, slot in enumerate(class_timetable[cls][day]):
                subj = slot['s']
                if subj in SPECIALS:
                    free_teachers = sorted(
                        [
                            {
                                'name': t,
                                'day_load': sum(
                                    1 for s in teacher_sched[t][day] if s is not None
                                ),
                                'wk_load': wk_load[t],
                            }
                            for t in all_teachers
                            if teacher_sched[t][day][pi] is None
                        ],
                        key=lambda x: (x['day_load'], x['wk_load']),
                    )
                    cov_list.append({
                        'period': PERIODS[pi],
                        'time':   TIMES[pi],
                        'type':   subj,
                        'free':   free_teachers,
                    })
            coverage[cls][day] = cov_list

    # teacher_load_detail — derive from teacher_view (subject → classes)
    from collections import defaultdict
    teacher_load_detail = {}
    for t in all_teachers:
        subj_cls = defaultdict(set)
        for day in DAYS:
            for slot in teacher_view[t][day]:
                if slot and slot['s'] not in ('', None) and slot['s'] not in SPECIALS:
                    subj_cls[slot['s']].add(slot['cls'])
        rows = []
        for subj in sorted(subj_cls):
            classes = sorted(subj_cls[subj])
            # count actual slots per week for this subject
            count = sum(
                1
                for day in DAYS
                for slot in teacher_view[t][day]
                if slot and slot['s'] == subj
            )
            per_class = count // len(classes) if classes else 0
            rows.append({
                'subject':   subj,
                'classes':   ', '.join(classes),
                'per_class': per_class,
                'subtotal':  count,
            })
        teacher_load_detail[t] = rows

    return (
        class_timetable, teacher_sched, teacher_view,
        coverage, wk_load, special_counts, all_teachers,
        teacher_load_detail,
    )

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = Path(__file__).parent / 'timetable_with_4_periods_v2.xlsx'
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {xlsx_path.name} …")
    wb = openpyxl.load_workbook(xlsx_path)

    sections     = parse_sections(wb)
    teacher_data = parse_teachers(wb)

    print(f"  Classes parsed : {sorted(sections.keys())}")
    print(f"  Teachers parsed: {sorted(teacher_data.keys())}")

    structs = build_structures(sections, teacher_data)

    # Patch CLASS_ORDER reference inside _build_html via module globals
    import html_exporter as hx2
    # _build_html uses json.dumps(CLASS_ORDER) — patch via the module
    import event_generator
    event_generator.CLASS_ORDER = CLASS_ORDER

    output = Path(__file__).parent / 'Timetable_Tools.html'
    html   = hx2._build_html(*structs)

    # Patch the classes_json inside the generated HTML so it uses our CLASS_ORDER
    # (event_generator.CLASS_ORDER was patched above so it should already be right)
    output.write_text(html, encoding='utf-8')
    size_kb = output.stat().st_size / 1024
    print(f"  HTML written → {output}  ({size_kb:.1f} KB)")

if __name__ == '__main__':
    main()
