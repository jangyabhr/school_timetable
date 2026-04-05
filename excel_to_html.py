# excel_to_html.py
#
# Read a (possibly user-modified) timetable.xlsx produced by main.py / exporter.py
# and regenerate Timetable_Tools.html without re-running the solver.
#
# Usage:
#   python excel_to_html.py [--input timetable.xlsx] [--output Timetable_Tools.html]

import argparse
import re
import sys
from collections import Counter

import openpyxl

from event_generator import CLASS_ORDER
from html_exporter import generate_html

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

NON_DATA_SHEETS = {"Dashboard", "Teacher Loads", "All Sections", "All Teachers"}

# exporter.py names class sheets "Class 6A", "Class 7B", etc.
CLASS_SHEET_PREFIX = "Class "
CLASS_SHEET_NAMES  = {f"Class {cls}": cls for cls in CLASS_ORDER}  # sheet_name → cls
CLASS_SHEET_SET    = set(CLASS_SHEET_NAMES.keys())

# Per-class sheet layout (written by exporter._write_timetable_grid):
#   Row 1  : merged title  — skip
#   Row 2  : header row    — skip
#   Rows 3–8 : Monday–Saturday (day index 0–5)
#   Cols 2–7 : P1–P6        (period index 0–5)
DATA_START_ROW = 3   # 1-based
DATA_START_COL = 2   # 1-based (col A = label, col B = P1, …)
NUM_DAYS    = 6
NUM_PERIODS = 6


# ---------------------------------------------------------------------------
# Cell-value parsers
# ---------------------------------------------------------------------------

_LAB_RE = re.compile(r'\s*\(Lab\)', re.IGNORECASE)


def _strip_lab(text):
    """Remove ' (Lab)' suffix; return (clean_text, is_lab)."""
    if _LAB_RE.search(text):
        return _LAB_RE.sub('', text).strip(), True
    return text.strip(), False


def _parse_class_cell(raw):
    """
    Parse a cell from a per-class sheet.

    Returns (subject, teacher_or_None, is_lab) or None if the cell is empty.

    Cell formats written by exporter.py:
      "Free"
      "Subject"
      "Subject (Lab)"
      "Subject\nTeacher"
      "Subject (Lab)\nTeacher"
    """
    if raw is None:
        return None
    val = str(raw).strip()
    if not val:
        return None

    lines = val.split('\n')
    first, rest = lines[0].strip(), lines[1].strip() if len(lines) > 1 else ''

    if first == 'Free':
        return ('Free', None, False)

    subject, is_lab = _strip_lab(first)
    teacher = rest if rest else None
    return (subject, teacher, is_lab)


def _parse_teacher_cell(raw):
    """
    Parse a cell from a per-teacher sheet.

    Returns (cls, subject, is_lab) or None if the cell is empty.

    Cell formats written by exporter.py:
      ""                        → None
      "Duty\n(6A)"              → ('6A', 'Free', False)
      "Subject\n(6A)"           → ('6A', 'Subject', False)
      "Subject (Lab)\n(6A)"     → ('6A', 'Subject', True)
    """
    if raw is None:
        return None
    val = str(raw).strip()
    if not val:
        return None

    lines = val.split('\n')
    first = lines[0].strip()
    second = lines[1].strip() if len(lines) > 1 else ''

    # Extract class name from "(6A)" format
    cls_match = re.match(r'^\((.+)\)$', second)
    cls = cls_match.group(1) if cls_match else second

    if first == 'Duty':
        return (cls, 'Free', False)

    subject, is_lab = _strip_lab(first)
    return (cls, subject, is_lab)


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _parse_class_sheet(ws):
    """
    Returns class_grid: list[6][6] of (subject, teacher_or_None, is_lab) or None.
    Indices: [day 0-5][period 0-5].
    """
    grid = [[None] * NUM_PERIODS for _ in range(NUM_DAYS)]
    for day_idx in range(NUM_DAYS):
        row = DATA_START_ROW + day_idx
        for period_idx in range(NUM_PERIODS):
            col = DATA_START_COL + period_idx
            cell = ws.cell(row=row, column=col)
            parsed = _parse_class_cell(cell.value)
            if parsed is not None:
                grid[day_idx][period_idx] = parsed
    return grid


def _parse_teacher_sheet(ws):
    """
    Returns teacher_grid: list[6][6] of (cls, subject, is_lab) or None.
    """
    grid = [[None] * NUM_PERIODS for _ in range(NUM_DAYS)]
    for day_idx in range(NUM_DAYS):
        row = DATA_START_ROW + day_idx
        for period_idx in range(NUM_PERIODS):
            col = DATA_START_COL + period_idx
            cell = ws.cell(row=row, column=col)
            parsed = _parse_teacher_cell(cell.value)
            if parsed is not None:
                grid[day_idx][period_idx] = parsed
    return grid


# ---------------------------------------------------------------------------
# Main conversion logic
# ---------------------------------------------------------------------------

def excel_to_timetable_state(wb):
    """
    Read an openpyxl Workbook and return (timetable_state, events).
    """
    sheet_names = wb.sheetnames

    # ── Identify sheet categories ────────────────────────────────────────────
    teacher_sheet_names = [
        name for name in sheet_names
        if name not in CLASS_SHEET_SET and name not in NON_DATA_SHEETS
    ]

    # ── Parse class sheets ───────────────────────────────────────────────────
    # class_data[cls][day][period] = (subject, teacher_or_None, is_lab) or None
    class_data = {}
    for cls in CLASS_ORDER:
        sheet_name = f"Class {cls}"
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: Sheet '{sheet_name}' not found in workbook — skipping.", file=sys.stderr)
            class_data[cls] = [[None] * NUM_PERIODS for _ in range(NUM_DAYS)]
        else:
            class_data[cls] = _parse_class_sheet(wb[sheet_name])

    # ── Parse teacher sheets ─────────────────────────────────────────────────
    # teacher_data[teacher_name][day][period] = (cls, subject, is_lab) or None
    teacher_data = {}
    for name in teacher_sheet_names:
        teacher_data[name] = _parse_teacher_sheet(wb[name])

    # ── Cross-reference: fill teacher for Free/Library cells ─────────────────
    # Build a lookup: (day, period, cls) → teacher_name
    duty_lookup = {}   # (day, period, cls) → teacher_name
    for teacher_name, grid in teacher_data.items():
        for day in range(NUM_DAYS):
            for period in range(NUM_PERIODS):
                entry = grid[day][period]
                if entry is not None:
                    cls_t, subj_t, _ = entry
                    key = (day, period, cls_t)
                    if key not in duty_lookup:
                        duty_lookup[key] = teacher_name

    for cls in CLASS_ORDER:
        for day in range(NUM_DAYS):
            for period in range(NUM_PERIODS):
                cell = class_data[cls][day][period]
                if cell is None:
                    continue
                subject, teacher, is_lab = cell
                if teacher is None:
                    # Try to find a teacher via teacher sheets
                    found = duty_lookup.get((day, period, cls))
                    if found:
                        class_data[cls][day][period] = (subject, found, is_lab)

    # ── Build timetable_state ────────────────────────────────────────────────
    timetable_state = {}
    idx = 0
    for cls_idx, cls in enumerate(CLASS_ORDER):
        for day in range(NUM_DAYS):
            for period in range(NUM_PERIODS):
                cell = class_data[cls][day][period]
                if cell is None:
                    continue
                subject, teacher, is_lab = cell
                timetable_state[idx] = {
                    'class':     cls,
                    'class_idx': cls_idx,
                    'day':       day,
                    'period':    period,
                    'subject':   subject,
                    'teacher':   teacher,
                    'is_lab':    is_lab,
                    'slot_id':   idx,
                }
                idx += 1

    # ── Derive events (weekly_load per teacher/subject/class) ────────────────
    counter = Counter(
        (p['teacher'], p['subject'], p['class'])
        for p in timetable_state.values()
        if p['teacher']
    )
    events = [
        {'teacher': t, 'subject': s, 'class': c, 'weekly_load': n}
        for (t, s, c), n in counter.items()
    ]

    return timetable_state, events


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Convert a (modified) timetable.xlsx into Timetable_Tools.html.'
    )
    parser.add_argument(
        '--input', '-i',
        default='timetable.xlsx',
        metavar='EXCEL_FILE',
        help='Path to the Excel workbook (default: timetable.xlsx)',
    )
    parser.add_argument(
        '--output', '-o',
        default='Timetable_Tools.html',
        metavar='HTML_FILE',
        help='Path for the generated HTML file (default: Timetable_Tools.html)',
    )
    args = parser.parse_args()

    print(f"Reading {args.input} …")
    try:
        wb = openpyxl.load_workbook(args.input, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: '{args.input}' not found.", file=sys.stderr)
        sys.exit(1)

    timetable_state, events = excel_to_timetable_state(wb)

    placements = len(timetable_state)
    teachers   = len({p['teacher'] for p in timetable_state.values() if p['teacher']})
    print(f"Parsed {placements} placements across {teachers} teachers.")

    generate_html(timetable_state, events, args.output)


if __name__ == '__main__':
    main()
