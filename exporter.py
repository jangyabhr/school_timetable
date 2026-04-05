# exporter.py

import openpyxl
from collections import defaultdict, Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from constraints import (
    DAYS_PER_WEEK,
    PERIODS_PER_DAY,
    PERIOD_NAMES_DISPLAY,
    PERIOD_TIMES,
)
from event_generator import CLASS_ORDER

# ---------------------------------------------------------------------------
# Colour palette  (openpyxl uses ARGB hex, no leading #)
# ---------------------------------------------------------------------------

# Per-subject pastel palette — one unique colour per subject
SUBJECT_COLOURS = {
    "Math":        "FFBBDEFB",  # pastel blue
    "English":     "FFB2EBF2",  # pastel cyan
    "Science":     "FFC8E6C9",  # pastel green
    "SST":         "FFFFF9C4",  # pastel yellow
    "Hindi":       "FFFFE0B2",  # pastel amber
    "Odia":        "FFFCE4EC",  # pastel pink
    "Sanskrit":    "FFE8EAF6",  # pastel indigo
    "CS":          "FFE0F2F1",  # pastel teal
    "IT":          "FFF3E5F5",  # pastel purple
    "Physics":     "FFE1F5FE",  # pastel light-blue
    "Chemistry":   "FFFBE9E7",  # pastel deep-orange
    "Biology":     "FFE8F5E9",  # pastel light-green
    "Library":     "FFF9FBE7",  # pastel lime
}

# Per-section pastel palette — used on teacher-wise sheets to colour by class
SECTION_COLOURS = {
    "6A":  "FFFADADD",  "6B":  "FFFDE8D8",
    "7A":  "FFD5F5E3",  "7B":  "FFD6EAF8",
    "8A":  "FFFFF3CD",  "8B":  "FFE8DAEF",
    "9A":  "FFDBEAFE",  "9B":  "FFD5D8DC",
    "10A": "FFFDEDEC",  "10B": "FFE9F7EF",
    "11":  "FFFEF9E7",  "12":  "FFEAF4FB",
}

COLOUR_FREE     = "FFD5D8DC"   # light grey-blue — Free (duty) periods
COLOUR_EMPTY    = "FFF5F5F5"   # light grey      — empty cell
COLOUR_HEADER   = "FF2E4057"   # dark blue       — header row/col
COLOUR_SUBHEAD  = "FF4A6FA5"   # medium blue     — sub-headers
COLOUR_WHITE    = "FFFFFFFF"
COLOUR_DASH_H   = "FF1B2631"   # near-black      — dashboard section headers

DAY_NAMES    = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIOD_NAMES = PERIOD_NAMES_DISPLAY   # ["Drill","P1","P2","Break","P3","P4","P5","P6"]


def _get_fill(subject):
    if subject is None:
        return PatternFill("solid", fgColor=COLOUR_EMPTY)
    if subject == "Free":
        return PatternFill("solid", fgColor=COLOUR_FREE)
    colour = SUBJECT_COLOURS.get(subject, COLOUR_WHITE)
    return PatternFill("solid", fgColor=colour)


def _get_section_fill(section):
    """Pastel fill keyed by class section — used on teacher-wise sheets."""
    colour = SECTION_COLOURS.get(section, COLOUR_WHITE)
    return PatternFill("solid", fgColor=colour)


def _thin_border():
    side = Side(style="thin", color="FFCCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _thick_border():
    side = Side(style="medium", color="FF888888")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor=COLOUR_HEADER)


def _subhead_fill():
    return PatternFill("solid", fgColor=COLOUR_SUBHEAD)


# ---------------------------------------------------------------------------
# Grid builder — days in rows, periods in columns
# grid[day][period] = (subject, teacher, is_lab)  or  None
# ---------------------------------------------------------------------------

def _build_class_grid(section, timetable_state):
    grid = [[None] * PERIODS_PER_DAY for _ in range(DAYS_PER_WEEK)]
    for key, placement in timetable_state.items():
        if placement["class"] != section:
            continue
        day     = placement["day"]
        period  = placement["period"]
        subject = placement["subject"]
        teacher = placement.get("teacher") or ""
        is_lab  = placement.get("is_lab", False)
        if not (0 <= day < DAYS_PER_WEEK and 0 <= period < PERIODS_PER_DAY):
            print(f"WARNING: {section} {subject} has out-of-range day={day} period={period}, skipping")
            continue
        grid[day][period] = (subject, teacher, is_lab)
    return grid


def _build_teacher_grid(teacher_name, timetable_state):
    """grid[day][period] = (subject, class_name, is_lab) or None"""
    grid = [[None] * PERIODS_PER_DAY for _ in range(DAYS_PER_WEEK)]
    for key, placement in timetable_state.items():
        if placement.get("teacher") != teacher_name:
            continue
        day     = placement["day"]
        period  = placement["period"]
        subject = placement["subject"]
        cls     = placement["class"]
        is_lab  = placement.get("is_lab", False)
        if not (0 <= day < DAYS_PER_WEEK and 0 <= period < PERIODS_PER_DAY):
            continue
        grid[day][period] = (subject, cls, is_lab)
    return grid


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_timetable_grid(ws, title, row_labels, col_labels, grid, cell_fill_fn, cell_value_fn, start_row=1):
    """
    Generic grid writer. Rows = days, cols = periods.
    row_labels : list of strings (one per data row)
    col_labels : list of strings (one per data column)
    grid       : grid[row][col] = cell_data or None
    cell_fill_fn(cell_data)  → PatternFill
    cell_value_fn(cell_data) → str
    Returns the next free row number.
    """
    num_cols = len(col_labels)

    # Title
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=num_cols + 1)
    tc = ws.cell(row=start_row, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color=COLOUR_WHITE)
    tc.fill      = _header_fill()
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22

    # Header row: "Day" + period names
    header_row = start_row + 1
    day_hdr = ws.cell(row=header_row, column=1, value="Day")
    day_hdr.fill      = _header_fill()
    day_hdr.font      = Font(bold=True, color=COLOUR_WHITE)
    day_hdr.alignment = Alignment(horizontal="center")

    for ci, col_label in enumerate(col_labels):
        # Show period name + time on two lines (e.g. "P1\n7:30–8:10")
        time_label = PERIOD_TIMES[ci] if ci < len(PERIOD_TIMES) else ""
        cell_val   = f"{col_label}\n{time_label}" if time_label else col_label
        cell = ws.cell(row=header_row, column=ci + 2, value=cell_val)
        cell.fill      = _header_fill()
        cell.font      = Font(bold=True, color=COLOUR_WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 30

    # Data rows
    for ri, row_label in enumerate(row_labels):
        data_row = header_row + 1 + ri
        ws.row_dimensions[data_row].height = 38

        lbl = ws.cell(row=data_row, column=1, value=row_label)
        lbl.fill      = _subhead_fill()
        lbl.font      = Font(bold=True, color=COLOUR_WHITE)
        lbl.alignment = Alignment(horizontal="center", vertical="center")

        for ci in range(num_cols):
            cell_data = grid[ri][ci]
            cell = ws.cell(row=data_row, column=ci + 2)
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.value     = cell_value_fn(cell_data)
            cell.fill      = cell_fill_fn(cell_data)

    # Column widths
    ws.column_dimensions["A"].width = 12
    for ci in range(num_cols):
        ws.column_dimensions[get_column_letter(ci + 2)].width = 16

    return header_row + 1 + len(row_labels)


def _write_class_sheet(ws, section, timetable_state):
    grid = _build_class_grid(section, timetable_state)

    def fill_fn(data):
        return _get_fill(data[0] if data else None)

    def val_fn(data):
        if not data:
            return "Free"
        subject, teacher, is_lab = data
        lab_suffix = " (Lab)" if is_lab else ""
        return f"{subject}{lab_suffix}\n{teacher}" if teacher else f"{subject}{lab_suffix}"

    _write_timetable_grid(
        ws,
        title      = f"Timetable — Class {section}",
        row_labels = DAY_NAMES,
        col_labels = PERIOD_NAMES,
        grid       = grid,
        cell_fill_fn  = fill_fn,
        cell_value_fn = val_fn,
    )


def _write_teacher_sheet(ws, teacher_name, timetable_state):
    grid = _build_teacher_grid(teacher_name, timetable_state)

    def fill_fn(data):
        if not data:
            return PatternFill("solid", fgColor=COLOUR_EMPTY)
        return _get_section_fill(data[1])   # data[1] = cls (section)

    def val_fn(data):
        if not data:
            return ""
        subject, cls, is_lab = data
        if subject == "Free":
            return f"Duty\n({cls})"
        lab_suffix = " (Lab)" if is_lab else ""
        return f"{subject}{lab_suffix}\n({cls})"

    _write_timetable_grid(
        ws,
        title      = f"Timetable — {teacher_name}",
        row_labels = DAY_NAMES,
        col_labels = PERIOD_NAMES,
        grid       = grid,
        cell_fill_fn  = fill_fn,
        cell_value_fn = val_fn,
    )


# ---------------------------------------------------------------------------
# Master aggregate sheets
# ---------------------------------------------------------------------------

def _write_master_sections_sheet(ws, timetable_state):
    """All 12 class timetables stacked vertically, one blank row between each."""
    next_row = 1
    for section in CLASS_ORDER:
        grid = _build_class_grid(section, timetable_state)

        def fill_fn(data):
            return _get_fill(data[0] if data else None)

        def val_fn(data):
            if not data:
                return "Free"
            subject, teacher, is_lab = data
            lab_suffix = " (Lab)" if is_lab else ""
            return f"{subject}{lab_suffix}\n{teacher}" if teacher else f"{subject}{lab_suffix}"

        next_row = _write_timetable_grid(
            ws,
            title         = f"Timetable — Class {section}",
            row_labels    = DAY_NAMES,
            col_labels    = PERIOD_NAMES,
            grid          = grid,
            cell_fill_fn  = fill_fn,
            cell_value_fn = val_fn,
            start_row     = next_row,
        )
        next_row += 1   # blank row between tables


def _write_master_teachers_sheet(ws, timetable_state, teachers):
    """All teacher timetables stacked vertically, one blank row between each."""
    next_row = 1
    for teacher_name in teachers:
        grid = _build_teacher_grid(teacher_name, timetable_state)

        def fill_fn(data):
            if not data:
                return PatternFill("solid", fgColor=COLOUR_EMPTY)
            return _get_section_fill(data[1])   # data[1] = cls (section)

        def val_fn(data):
            if not data:
                return ""
            subject, cls, is_lab = data
            if subject == "Free":
                return f"Duty\n({cls})"
            lab_suffix = " (Lab)" if is_lab else ""
            return f"{subject}{lab_suffix}\n({cls})"

        next_row = _write_timetable_grid(
            ws,
            title         = f"Timetable — {teacher_name}",
            row_labels    = DAY_NAMES,
            col_labels    = PERIOD_NAMES,
            grid          = grid,
            cell_fill_fn  = fill_fn,
            cell_value_fn = val_fn,
            start_row     = next_row,
        )
        next_row += 1   # blank row between tables


# ---------------------------------------------------------------------------
# Violations section (written into the Dashboard)
# ---------------------------------------------------------------------------

COLOUR_VIOL_HDR  = "FFC0392B"   # red      — violations section header
COLOUR_VIOL_ROW  = "FFFCE4D6"   # salmon   — violation row cells
COLOUR_OK_ROW    = "FFE9F7EF"   # pale green — all-clear row


def _write_violations_section(ws, violations, start_row):
    """
    Writes a Validation Report block starting at start_row, columns A–C.
    Each violation row shows: # | Violation | Suggested Fix
    """
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    bold_white = Font(bold=True, color=COLOUR_WHITE)

    if not violations:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row,   end_column=3)
        hdr = ws.cell(row=start_row, column=1,
                      value="Validation Report — All checks passed")
        hdr.fill      = PatternFill("solid", fgColor=COLOUR_DASH_H)
        hdr.font      = Font(bold=True, size=12, color=COLOUR_WHITE)
        hdr.alignment = center
        ws.row_dimensions[start_row].height = 20

        ok_row = start_row + 1
        ws.merge_cells(start_row=ok_row, start_column=1,
                       end_row=ok_row,   end_column=3)
        ok = ws.cell(row=ok_row, column=1, value="No violations found")
        ok.fill      = PatternFill("solid", fgColor=COLOUR_OK_ROW)
        ok.font      = Font(bold=True, color="FF1E8449")
        ok.alignment = center
        ok.border    = _thin_border()
        ws.row_dimensions[ok_row].height = 18
        return

    # Header
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=3)
    hdr = ws.cell(row=start_row, column=1,
                  value=f"Validation Report — {len(violations)} Violation(s) Found")
    hdr.fill      = PatternFill("solid", fgColor=COLOUR_VIOL_HDR)
    hdr.font      = Font(bold=True, size=12, color=COLOUR_WHITE)
    hdr.alignment = center
    ws.row_dimensions[start_row].height = 20

    # Column headers
    col_row = start_row + 1
    for col, label in enumerate(["#", "Violation", "Suggested Fix"], start=1):
        c = ws.cell(row=col_row, column=col, value=label)
        c.fill      = _subhead_fill()
        c.font      = bold_white
        c.alignment = center
        c.border    = _thin_border()
    ws.row_dimensions[col_row].height = 16

    # Violation rows
    for i, v in enumerate(violations, start=1):
        row = col_row + i
        ws.row_dimensions[row].height = 42

        num_c = ws.cell(row=row, column=1, value=i)
        num_c.fill      = PatternFill("solid", fgColor=COLOUR_VIOL_ROW)
        num_c.border    = _thin_border()
        num_c.alignment = center
        num_c.font      = Font(bold=True)

        msg_c = ws.cell(row=row, column=2, value=v["message"])
        msg_c.fill      = PatternFill("solid", fgColor=COLOUR_VIOL_ROW)
        msg_c.border    = _thin_border()
        msg_c.alignment = left

        sug_c = ws.cell(row=row, column=3, value=v["suggestion"])
        sug_c.fill      = PatternFill("solid", fgColor=COLOUR_VIOL_ROW)
        sug_c.border    = _thin_border()
        sug_c.alignment = left

    # Widen columns B and C to accommodate violation text
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55


# ---------------------------------------------------------------------------
# Dashboard sheet
# ---------------------------------------------------------------------------

def _write_dashboard(ws, timetable_state, events, violations=None):
    """
    Section 1: Teacher Load Summary — teacher → total periods/week
    Section 2: Weekly periods per subject per section (section × subject matrix)
    """
    bold_white  = Font(bold=True, color=COLOUR_WHITE)
    bold_dark   = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")

    def _sec_header(row, col, text, width_cols):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + width_cols - 1)
        c = ws.cell(row=row, column=col, value=text)
        c.fill      = PatternFill("solid", fgColor=COLOUR_DASH_H)
        c.font      = Font(bold=True, size=12, color=COLOUR_WHITE)
        c.alignment = center
        ws.row_dimensions[row].height = 20

    # ── Collect teacher load ─────────────────────────────────────────────────
    teacher_load = Counter()
    for placement in timetable_state.values():
        t = placement.get("teacher")
        if t:
            teacher_load[t] += 1

    sorted_teachers = sorted(teacher_load.keys())

    # ── Collect subject counts per section ──────────────────────────────────
    section_subject_counts = defaultdict(Counter)
    for placement in timetable_state.values():
        section_subject_counts[placement["class"]][placement["subject"]] += 1

    # Determine all subjects (sorted)
    all_subjects = sorted({placement["subject"] for placement in timetable_state.values()})

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1 — Teacher Load Summary  (starts at row 1)
    # ════════════════════════════════════════════════════════════════════════
    _sec_header(1, 1, "Teacher Load Summary (periods / week)", 3)

    # Column headers
    for col, hdr in enumerate(["Teacher", "Periods / Week"], start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = _subhead_fill()
        c.font = bold_white
        c.alignment = center
        c.border = _thin_border()

    for ri, teacher in enumerate(sorted_teachers, start=3):
        ws.cell(row=ri, column=1, value=teacher).border = _thin_border()
        ws.cell(row=ri, column=1).alignment = left
        load_cell = ws.cell(row=ri, column=2, value=teacher_load[teacher])
        load_cell.border    = _thin_border()
        load_cell.alignment = center
        ws.row_dimensions[ri].height = 16

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    load_end_row = 2 + len(sorted_teachers)

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2 — Weekly periods per subject per section
    # (placed to the right of section 1, starting at column 4)
    # ════════════════════════════════════════════════════════════════════════
    LEFT_COL = 4   # sections start here
    num_subjects = len(all_subjects)

    _sec_header(1, LEFT_COL, "Weekly Periods per Subject per Section",
                num_subjects + 1)

    # Row header: "Section" + subject names
    ws.cell(row=2, column=LEFT_COL, value="Section").fill = _subhead_fill()
    ws.cell(row=2, column=LEFT_COL).font      = bold_white
    ws.cell(row=2, column=LEFT_COL).alignment = center
    ws.cell(row=2, column=LEFT_COL).border    = _thin_border()

    for ci, subj in enumerate(all_subjects):
        c = ws.cell(row=2, column=LEFT_COL + 1 + ci, value=subj)
        c.fill      = _subhead_fill()
        c.font      = bold_white
        c.alignment = center
        c.border    = _thin_border()

    for ri, section in enumerate(CLASS_ORDER, start=3):
        sec_cell = ws.cell(row=ri, column=LEFT_COL, value=section)
        sec_cell.fill      = PatternFill("solid", fgColor=COLOUR_HEADER)
        sec_cell.font      = bold_white
        sec_cell.alignment = center
        sec_cell.border    = _thin_border()
        ws.row_dimensions[ri].height = 16

        for ci, subj in enumerate(all_subjects):
            count = section_subject_counts[section].get(subj, 0)
            c = ws.cell(row=ri, column=LEFT_COL + 1 + ci,
                        value=count if count > 0 else "")
            c.border    = _thin_border()
            c.alignment = center
            if count > 0:
                c.fill = _get_fill(subj)

    # Column widths for subject matrix
    ws.column_dimensions[get_column_letter(LEFT_COL)].width = 12
    for ci in range(num_subjects):
        ws.column_dimensions[get_column_letter(LEFT_COL + 1 + ci)].width = 11

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3 — Validation Report  (below Teacher Load, columns A–C)
    # ════════════════════════════════════════════════════════════════════════
    violations_start = load_end_row + 2
    _write_violations_section(ws, violations or [], violations_start)


# ---------------------------------------------------------------------------
# Validation check before export
# ---------------------------------------------------------------------------

def validate_before_export(timetable_state, events):
    """
    Returns a list of violation dicts, each with keys:
      "message"    — human-readable description of the problem
      "suggestion" — recommended corrective action
    """
    violations = []
    seen_teacher = {}
    seen_class   = {}

    for key, p in timetable_state.items():
        day, period = p["day"], p["period"]
        subject = p.get("subject", "?")
        cls     = p.get("class", "?")
        day_label    = DAY_NAMES[day] if 0 <= day < len(DAY_NAMES) else f"Day{day}"
        period_label = PERIOD_NAMES[period] if 0 <= period < len(PERIOD_NAMES) else f"P{period}"

        teacher = p.get("teacher")
        if teacher:
            tk = (teacher, day, period)
            if tk in seen_teacher:
                other_p = timetable_state[seen_teacher[tk]]
                violations.append({
                    "message": (
                        f"Teacher clash: {teacher} is double-booked — "
                        f"{cls} {subject} and {other_p['class']} {other_p['subject']} "
                        f"both at {day_label} {period_label}"
                    ),
                    "suggestion": (
                        f"Move {cls} {subject} or {other_p['class']} {other_p['subject']} "
                        f"to any free period for {teacher}"
                    ),
                })
            else:
                seen_teacher[tk] = key

        ck = (cls, day, period)
        if ck in seen_class:
            other_p = timetable_state[seen_class[ck]]
            violations.append({
                "message": (
                    f"Class clash: {cls} has {subject} and {other_p['subject']} "
                    f"both at {day_label} {period_label}"
                ),
                "suggestion": (
                    f"Move {subject} or {other_p['subject']} for {cls} "
                    f"to a different day or period"
                ),
            })
        else:
            seen_class[ck] = key

    placement_counts = Counter()
    for (event_idx, instance), _ in timetable_state.items():
        placement_counts[event_idx] += 1

    for event_idx, event in enumerate(events):
        placed   = placement_counts.get(event_idx, 0)
        expected = event["weekly_load"]
        if placed != expected:
            if placed < expected:
                suggestion = (
                    f"Re-run solver or manually add {expected - placed} more "
                    f"period(s) of {event['subject']} for {event['class']}"
                )
            else:
                suggestion = (
                    f"Remove {placed - expected} extra period(s) of "
                    f"{event['subject']} for {event['class']} from the timetable"
                )
            violations.append({
                "message": (
                    f"Load mismatch: {event['class']} {event['subject']} "
                    f"— expected {expected} period(s)/week, placed {placed}"
                ),
                "suggestion": suggestion,
            })

    return violations


# ---------------------------------------------------------------------------
# Teacher Load Sheet
# ---------------------------------------------------------------------------

def _write_teacher_loads_sheet(ws, events):
    """
    Writes a Teacher Load reference sheet.

    Columns: Teacher | Subject | Classes | Periods/Class | Subtotal
    One block per teacher (sorted), followed by a bold TOTAL row.
    Data is derived entirely from the events list produced by event_generator.
    """
    bold_white = Font(bold=True, color=COLOUR_WHITE)
    bold_dark  = Font(bold=True)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="Teacher Subject Load — Periods per Week")
    title.font      = Font(bold=True, size=13, color=COLOUR_WHITE)
    title.fill      = _header_fill()
    title.alignment = center
    ws.row_dimensions[1].height = 22

    # ── Column headers ─────────────────────────────────────────────────────────
    for col, hdr in enumerate(["Teacher", "Subject", "Classes", "Periods/Class", "Subtotal"], 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill      = _subhead_fill()
        c.font      = bold_white
        c.alignment = center
        c.border    = _thin_border()
    ws.row_dimensions[2].height = 16

    # ── Build data: teacher → [(subject, [classes], periods_per_class)] ────────
    # Group events by (teacher, subject, weekly_load) → list of classes
    from collections import defaultdict
    teacher_data = defaultdict(lambda: defaultdict(list))  # [teacher][subject+load] → classes

    for ev in events:
        teacher = ev.get("teacher")
        if not teacher:
            continue
        key = (ev["subject"], ev["weekly_load"])
        teacher_data[teacher][key].append(ev["class"])

    # Sort teachers; within each teacher sort by subject
    sorted_teachers = sorted(teacher_data.keys())

    row = 3
    for teacher in sorted_teachers:
        subj_map = teacher_data[teacher]
        teacher_total = 0

        # Sort rows: by subject name, then load descending
        sorted_rows = sorted(subj_map.items(), key=lambda x: (x[0][0], -x[0][1]))

        first_row_for_teacher = row
        for (subject, load), classes in sorted_rows:
            subtotal = load * len(classes)
            teacher_total += subtotal

            classes_str = ", ".join(sorted(classes))
            fill = _get_fill(subject)

            ws.cell(row=row, column=1, value=teacher).border  = _thin_border()
            ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=2, value=subject).fill    = fill
            ws.cell(row=row, column=2).border    = _thin_border()
            ws.cell(row=row, column=2).alignment = left
            ws.cell(row=row, column=3, value=classes_str).border = _thin_border()
            ws.cell(row=row, column=3).alignment = left
            ws.cell(row=row, column=4, value=load).border     = _thin_border()
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=5, value=subtotal).border = _thin_border()
            ws.cell(row=row, column=5).alignment = center
            ws.row_dimensions[row].height = 16
            row += 1

        # Total row for this teacher
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        total_lbl = ws.cell(row=row, column=1,
                            value=f"{teacher}  — TOTAL")
        total_lbl.fill      = PatternFill("solid", fgColor=COLOUR_HEADER)
        total_lbl.font      = Font(bold=True, color=COLOUR_WHITE)
        total_lbl.alignment = Alignment(horizontal="right", vertical="center")
        total_lbl.border    = _thin_border()

        total_val = ws.cell(row=row, column=5, value=teacher_total)
        total_val.fill      = PatternFill("solid", fgColor=COLOUR_HEADER)
        total_val.font      = Font(bold=True, color=COLOUR_WHITE)
        total_val.alignment = center
        total_val.border    = _thin_border()
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10


# ---------------------------------------------------------------------------
# Main Export Entry Point
# ---------------------------------------------------------------------------

def export_timetable(timetable_state, events, output_path="timetable.xlsx"):
    print("── Running pre-export validation ──")
    unknown_classes = {p["class"] for p in timetable_state.values()} - set(CLASS_ORDER)
    if unknown_classes:
        print(f"WARNING: classes {sorted(unknown_classes)} not in CLASS_ORDER — skipped")
    violations = validate_before_export(timetable_state, events)

    if violations:
        print(f"   EXPORT WARNING — {len(violations)} violation(s) (see Dashboard sheet):")
        for v in violations:
            print(f"   ⚠  {v['message']}")
            print(f"      → {v['suggestion']}")
    else:
        print("   Validation passed — no violations found.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    teachers = sorted({p["teacher"] for p in timetable_state.values() if p.get("teacher")})

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────────
    ws_dash = wb.create_sheet(title="Dashboard")
    _write_dashboard(ws_dash, timetable_state, events, violations)

    # ── Sheet 2: Teacher Loads ───────────────────────────────────────────────
    ws_loads = wb.create_sheet(title="Teacher Loads")
    _write_teacher_loads_sheet(ws_loads, events)

    # ── Sheet 3: All Sections — every class timetable stacked ───────────────
    ws_all_sec = wb.create_sheet(title="All Sections")
    _write_master_sections_sheet(ws_all_sec, timetable_state)

    # ── Sheet 3: All Teachers — every teacher timetable stacked ─────────────
    ws_all_tch = wb.create_sheet(title="All Teachers")
    _write_master_teachers_sheet(ws_all_tch, timetable_state, teachers)

    # ── Sheets 4+: Class-wise timetables (one sheet per class) ──────────────
    for section in CLASS_ORDER:
        ws = wb.create_sheet(title=f"Class {section}")
        _write_class_sheet(ws, section, timetable_state)

    # ── Sheets N+: Teacher-wise timetables (one sheet per teacher) ───────────
    for teacher in teachers:
        ws = wb.create_sheet(title=teacher[:31])   # Excel sheet name limit = 31 chars
        _write_teacher_sheet(ws, teacher, timetable_state)

    wb.save(output_path)
    print(f"── Timetable saved to: {output_path} ──")
